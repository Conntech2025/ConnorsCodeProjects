import tkinter as tk
from tkinter import messagebox
import openpyxl
import os
import re

# Ensure the directory exists
folder_path = r"c:\apps"
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# Excel file path
file_name = os.path.join(folder_path, "loan_applications.xlsx")

# Create the Excel file if it doesn't exist
if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "General Information"
        ws.append(["First Name", "Last Name", "Email", "Phone", "Loan Amount", "Loan Term", "Employment Status", "Annual Income"])
        wb.create_sheet("Work History").append(["Current Employer", "Years at Job", "Previous Employer 1", "Years", "Previous Employer 2", "Years", "Previous Employer 3", "Years"])
        wb.create_sheet("Education History").append(["High School", "Graduation Year", "College 1", "Degree 1", "College 2", "Degree 2", "College 3", "Degree 3", "College 4", "Degree 4"])
        wb.save(file_name)
else:
        wb = openpyxl.load_workbook(file_name)
        sheets_needed = {
            "General Information": ["First Name", "Last Name", "Email", "Phone", "Loan Amount", "Loan Term", "Employment Status", "Annual Income"],
            "Work History": ["Current Employer", "Years at Job", "Previous Employer 1", "Years", "Previous Employer 2", "Years", "Previous Employer 3", "Years"],
            "Education History": ["High School", "Graduation Year", "College 1", "Degree 1", "College 2", "Degree 2", "College 3", "Degree 3", "College 4", "Degree 4"]
        }
        for sheet_name, headers in sheets_needed.items():
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name).append(headers)
        wb.save(file_name)

# Function to validate email
def is_valid_email(email):
    pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.match(pattern, email)

# Function to validate phone number (eg., (321)-456-7869)
def is_valid_phone(phone):
    pattern = r"^\(\d{3}\)-\d{3}-\d{4}$"
    return re.match(pattern, phone)

# Function to validate numeric fields
def is_valid_number(value):
    try:
        float_val = float(value)
        return round(float_val, 2) == float_val
    except ValueError:
        return False

# Function to validate graduation year
def is_valid_year(year):
    return year.isdigit() and 1900 <= int(year) <= 2025

# Auto-format phone number as user types
def format_phone(event):
    text = entry_phone.get().replace("(", "").replace(")", "").replace("-", "")
    if text.isdigit():
        if len(text) > 0:
            formatted = "(" + text[:3]
            if len(text) >= 3:
                formatted += ")-" + text[3:6]
            if len(text) >= 6:
                formatted += "-" + text[6:10]
            entry_phone.delete(0, tk.END)
            entry_phone.insert(0, formatted)

# Function to switch frames
def show_frame(frame):
    frame.tkraise()


def save_data():
    # General Info
    first_name = entry_first_name.get().strip()
    last_name = entry_last_name.get().strip()
    email = entry_email.get().strip()
    phone = entry_phone.get().strip()
    loan_amount = entry_loan_amount.get().strip()
    loan_term = loan_term_var.get()
    employment_status = employment_var.get()
    annual_income = entry_income.get().strip()

    # Work History
    current_employer = entry_current_employer.get()
    years_current = entry_years_current.get().strip()
    prev_employers = []
    prev_employer_1 = entry_prev_emp1.get().strip()
    years_emp1 = entry_prev_years1.get().strip()
    prev_employers.append((prev_employer_1, years_emp1))
    prev_employer_2 = entry_prev_emp2.get().strip()
    years_emp2 = entry_prev_years2.get().strip()
    prev_employers.append((prev_employer_2, years_emp2))
    prev_employer_3 = entry_prev_emp3.get().strip()
    years_emp3 = entry_prev_years3.get().strip()
    prev_employers.append((prev_employer_3, years_emp3))

    # Education History
    high_school = entry_high_school_name.get().strip()
    grad_year = entry_graduation_year.get().strip()
    college_1 = entry_college_name.get().strip()
    degree_1 = degree_earned_var.get().strip()
    degree_2 = degree_2_var.get().strip()
    degree_3 = degree_3_var.get().strip()
    college_2 = entry_college_2.get().strip()
    college_3 = entry_college_3.get().strip()
    college_4 = entry_college_4.get().strip()
    degree_4 = degree_4_var.get().strip()

    try:
        wb = openpyxl.load_workbook(file_name)
        ws1 = wb["General Information"]
        ws2 = wb["Work History"]
        ws3 = wb["Education History"]
        ws1.append([first_name, last_name, email, phone, loan_amount, loan_term, employment_status, annual_income])
        work_row = [current_employer, years_current]
        for emp, yrs in prev_employers:
            work_row.extend([emp, yrs])
        ws2.append(work_row)
        ws3.append([high_school, grad_year, college_1, degree_1, college_2, degree_2, college_3, degree_3, college_4, degree_4])
        wb.save(file_name)
        wb.close()
        messagebox.showinfo("Success", "Application Submitted Successfully!")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data: {e}")



def next_button1():
    # General Info
    first_name = entry_first_name.get().strip()
    last_name = entry_last_name.get().strip()
    email = entry_email.get().strip()
    phone = entry_phone.get().strip()
    loan_amount = entry_loan_amount.get().strip()
    loan_term = loan_term_var.get()
    employment_status = employment_var.get()
    annual_income = entry_income.get().strip()

    # Validations
    if not first_name:
        messagebox.showerror("Error", "First Name is required.")
        return
    
    if not last_name:
        messagebox.showerror("Error", "Last Name is required.")
        return
    
    if len(first_name) > 25:
        messagebox.showerror("Error", "Fast Name cannot exceed 25 characters.")
        return

    if len(last_name) > 35:
        messagebox.showerror("Error", "Last Name cannot exceed 35 characters.")
        return
    
    if not is_valid_email(email):
        messagebox.showerror("Error", "Please enter a valid email address.")
        return
    
    if not is_valid_phone(phone):
        messagebox.showerror("Error", "Please enter a valid phone number (321)-456-7869.")
        return
    
    if not is_valid_number(loan_amount):
        messagebox.showerror("Error", "Loan Amount must be a numeric value.")
        return
    
    if not is_valid_number(annual_income):
        messagebox.showerror("Error", "Annual Income must be a numeric value.")
        return
    
    if loan_term == "Select One":
        messagebox.showerror("Error", "Please select a valid loan term.")
        return
    
    if employment_status == "Select One":
        messagebox.showerror("Error", "Please select a valid employment option.")
        return
    show_frame(frame2)

def next_button2():
    # Work History
    current_employer = entry_current_employer.get()
    years_current = entry_years_current.get().strip()
    prev_employers = []
    prev_employer_1 = entry_prev_emp1.get().strip()
    years_emp1 = entry_prev_years1.get().strip()
    prev_employers.append((prev_employer_1, years_emp1))
    prev_employer_2 = entry_prev_emp2.get().strip()
    years_emp2 = entry_prev_years2.get().strip()
    prev_employers.append((prev_employer_2, years_emp2))
    prev_employer_3 = entry_prev_emp3.get().strip()
    years_emp3 = entry_prev_years3.get().strip()
    prev_employers.append((prev_employer_3, years_emp3))
    
    # Validations
    if not current_employer or not is_valid_number(years_current):
        messagebox.showerror("Error", "Current Employer and Years must be valid.")
        return
    
    for emp, yrs in prev_employers:
        if emp and not yrs:
            messagebox.showerror("Error", "Previous Employer field requires Years.")
            return
        if yrs and not is_valid_number(yrs):
            messagebox.showerror("Error", "Years at previous jobs must be numeric.")
            return
    show_frame(frame3)

def submit_button():
    # Education History
    high_school = entry_high_school_name.get().strip()
    grad_year = entry_graduation_year.get().strip()
    college_1 = entry_college_name.get().strip()
    degree_1 = degree_earned_var.get().strip()
    degree_2 = degree_2_var.get().strip()
    degree_3 = degree_3_var.get().strip()
    college_2 = entry_college_2.get().strip()
    college_3 = entry_college_3.get().strip()
    college_4 = entry_college_4.get().strip()
    degree_4 = degree_4_var.get().strip()

    if not high_school:
        messagebox.showerror("Error", "High School Name is required.")
        return
        
    if not high_school or not is_valid_year(grad_year):
        messagebox.showerror("Error", "Graduation year must be between 1900 and 2025.")
        return
    
    if any(college and not degree for college, degree in [(college_1, degree_1), (college_2, degree_2), (college_3, degree_3), (college_4, degree_4)
    ]):
        messagebox.showerror("Error", "Degree is required if college is filled.")
        return
    save_data()
    

# Function to switch frames
def show_frame(frame):
    frame.tkraise()

# Create main application window
root = tk.Tk()
root.title("Acme Home Mortgage Application")
root.geometry("850x400")

# Create container for frames
container = tk.Frame(root)
container.place(relx=0.5, rely=0.1, anchor="n")  

# Frame 1
frame1 = tk.Frame(container)
frame1.grid(row=0, column=0, sticky="nsew")

title_label = tk.Label(frame1, text="Acme Home Mortgage Application", font=("Arial", 14, "bold"))
title_label.grid(row=0, column=0, columnspan=4, pady=10)

tk.Label(frame1, text="First Name:").grid(row=1, column=0, sticky="e", padx=5, pady=2)
entry_first_name = tk.Entry(frame1, width=25)
entry_first_name.grid(row=1, column=1, sticky="w")

tk.Label(frame1, text="Last Name:").grid(row=1, column=2, sticky="e", padx=5, pady=2)
entry_last_name = tk.Entry(frame1, width=35)
entry_last_name.grid(row=1, column=3, sticky="w")

tk.Label(frame1, text="Email:").grid(row=2, column=0, sticky="e", padx=5, pady=2)
entry_email = tk.Entry(frame1, width=25)
entry_email.grid(row=2, column=1, sticky="w")

tk.Label(frame1, text="Phone:").grid(row=2, column=2, sticky="e", padx=5, pady=2)
entry_phone = tk.Entry(frame1, width=25)
entry_phone.grid(row=2, column=3, sticky="w")

tk.Label(frame1, text="Loan Amount:").grid(row=3, column=0, sticky="e", padx=5, pady=2)
entry_loan_amount = tk.Entry(frame1, width=25)
entry_loan_amount.grid(row=3, column=1, sticky="w")

# Employment Status and Annual Income
tk.Label(frame1, text="Employment Status:").grid(row=4, column=0, sticky="e", padx=10, pady=5)
employment_var = tk.StringVar(value="Select One")
employment_options = ["Select One", "Employed", "Self-Employed", "Unemployed", "Student", "Retired"]
employment_menu = tk.OptionMenu(frame1, employment_var, *employment_options)
employment_menu.grid(row=4, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame1, text="Annual Income:").grid(row=5, column=0, sticky="e", padx=10, pady=5)
entry_income = tk.Entry(frame1, width=25)
entry_income.grid(row=5, column=1, padx=5, pady=5, sticky="w")

# Dropdowns
tk.Label(frame1, text="Loan Term:").grid(row=3, column=2, sticky="e", padx=10, pady=5)
loan_term_var = tk.StringVar(value="Select One")
loan_term_options = ["Select One", "10-year", "15-year", "30-year"]
loan_term_menu = tk.OptionMenu(frame1, loan_term_var, *loan_term_options)
loan_term_menu.grid(row=3, column=3, padx=5, pady=5, sticky="w")

# Frame 2
frame2 = tk.Frame(container)
frame2.grid(row=0, column=0, sticky="nsew")

tk.Label(frame2, text="Work History", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=4, pady=10)

tk.Label(frame2, text="Current Employer:").grid(row=1, column=0, sticky="e", padx=10, pady=5)
entry_current_employer = tk.Entry(frame2, width=25)
entry_current_employer.grid(row=1, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame2, text="Years at Job:").grid(row=1, column=2, sticky="e", padx=10, pady=5)
entry_years_current = tk.Entry(frame2, width=10)
entry_years_current.grid(row=1, column=3, padx=5, pady=5, sticky="w")

for i in range(3):
  tk.Label(frame2, text="Previous Employer:").grid(row=i+2, column=0, sticky="e", padx=10, pady=5)
  tk.Entry(frame2, width=25).grid(row=i+2, column=1, padx=5, pady=5, sticky="w")
  tk.Label(frame2, text="Years at Job:").grid(row=i+2, column=2, sticky="e", padx=10, pady=5)
  tk.Entry(frame2, width=10).grid(row=i+2, column=3, padx=5, pady=5, sticky="w")

# Previous Employers Section in Frame 2 
entry_prev_emp1 = tk.Entry(frame2, width=25)
entry_prev_emp1.grid(row=2, column=1, padx=5, pady=5, sticky="w")
entry_prev_years1 = tk.Entry(frame2, width=10)
entry_prev_years1.grid(row=2, column=3, padx=5, pady=5, sticky="w")

entry_prev_emp2 = tk.Entry(frame2, width=25)
entry_prev_emp2.grid(row=3, column=1, padx=5, pady=5, sticky="w")
entry_prev_years2 = tk.Entry(frame2, width=10)
entry_prev_years2.grid(row=3, column=3, padx=5, pady=5, sticky="w")

entry_prev_emp3 = tk.Entry(frame2, width=25)
entry_prev_emp3.grid(row=4, column=1, padx=5, pady=5, sticky="w")
entry_prev_years3 = tk.Entry(frame2, width=10)
entry_prev_years3.grid(row=4, column=3, padx=5, pady=5, sticky="w")

# Frame 3
frame3 = tk.Frame(container)
frame3.grid(row=0, column=0, sticky="nsew")

tk.Label(frame3, text="Education History", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=4, pady=10)

tk.Label(frame3, text="High School Name:").grid(row=1, column=0, sticky="e", padx=10, pady=5)
entry_high_school_name = tk.Entry(frame3, width=25)
entry_high_school_name.grid(row=1, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame3, text="Graduation Year:").grid(row=1, column=2, sticky="e", padx=10, pady=5)
entry_graduation_year = tk.Entry(frame3, width=10)
entry_graduation_year.grid(row=1, column=3, padx=5, pady=5, sticky="w")

# Degree dropdown options
DEGREE_OPTIONS = ["Select One", "AAS", "BS", "MS", "PhD", "EdD"]

# 1st College Attended
tk.Label(frame3, text="College Attended:").grid(row=2, column=0, sticky="e", padx=10, pady=5)
entry_college_name = tk.Entry(frame3, width=25)
entry_college_name.grid(row=2, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame3, text="Degree Earned:").grid(row=2, column=2, sticky="e", padx=10, pady=5)
degree_earned_var = tk.StringVar(value="Select One")
entry_degree_earned = tk.OptionMenu(frame3, degree_earned_var, *DEGREE_OPTIONS)
entry_degree_earned.grid(row=2, column=3, padx=5, pady=5, sticky="w")

# 2nd College Attended
tk.Label(frame3, text="2nd Previous College Attended:").grid(row=3, column=0, sticky="e", padx=10, pady=5)
entry_college_2 = tk.Entry(frame3, width=25)
entry_college_2.grid(row=3, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame3, text="Degree Earned:").grid(row=3, column=2, sticky="e", padx=10, pady=5)
degree_2_var = tk.StringVar(value="Select One")
entry_degree_2 = tk.OptionMenu(frame3, degree_2_var, *DEGREE_OPTIONS)
entry_degree_2.grid(row=3, column=3, padx=5, pady=5, sticky="w")

# 3rd College Attended
tk.Label(frame3, text="3rd Previous College Attended:").grid(row=4, column=0, sticky="e", padx=10, pady=5)
entry_college_3 = tk.Entry(frame3, width=25)
entry_college_3.grid(row=4, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame3, text="Degree Earned:").grid(row=4, column=2, sticky="e", padx=10, pady=5)
degree_3_var = tk.StringVar(value="Select One")
entry_degree_3 = tk.OptionMenu(frame3, degree_3_var, *DEGREE_OPTIONS)
entry_degree_3.grid(row=4, column=3, padx=5, pady=5, sticky="w")

# 4th College Attended
tk.Label(frame3, text="4th Previous College Attended:").grid(row=5, column=0, sticky="e", padx=10, pady=5)
entry_college_4 = tk.Entry(frame3, width=25, justify="center")
entry_college_4.grid(row=5, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame3, text="Degree Earned:").grid(row=5, column=2, sticky="e", padx=10, pady=5)
degree_4_var = tk.StringVar(value="Select One")
entry_degree_4 = tk.OptionMenu(frame3, degree_4_var, *DEGREE_OPTIONS)
entry_degree_4.grid(row=5, column=3, padx=5, pady=5, sticky="w")

# Back and Submit buttons (Frame 3)
button_frame = tk.Frame(frame3)
button_frame.grid(row=6, column=0, columnspan=4, pady=10)

back_button = tk.Button(button_frame, text="← Back", bg="gray", fg="white", font=("Arial", 12), command=lambda: show_frame(frame2))
back_button.pack(side="left", padx=5)

submit_button = tk.Button(button_frame, text="Submit Application", command=submit_button, bg="green", fg="white", font=("Arial", 12, "bold"))
submit_button.pack(side="left", padx=5)

# Back and Submit buttons (Frame 2)
button_frame = tk.Frame(frame2)
button_frame.grid(row=5, column=0, columnspan=4, pady=10)

back_button = tk.Button(button_frame, text="← Back", bg="gray", fg="white", font=("Arial", 12), command=lambda: show_frame(frame1))
back_button.pack(side="left", padx=5)

next_button = tk.Button(button_frame, text="Next →", bg="blue", fg="white", font=("Arial", 12), command=next_button2)
next_button.pack(side="right", padx=5)

# Back and Submit buttons (Frame 1)
next_button = tk.Button(frame1, text="Next →", bg="blue", fg="white", font=("Arial", 12), command=next_button1)
next_button.grid(row=6, column=3, pady=20, sticky="e", padx=10)

# Set the initial frame
show_frame(frame1)

# Run the Tkinter event loop
root.mainloop()
