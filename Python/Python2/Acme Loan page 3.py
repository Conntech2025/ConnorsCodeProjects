import tkinter as tk
from tkinter import messagebox
import openpyxl
import os
import re

# Ensure the directory exists
folder_path = r"c:\apps"
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# Excel file path
file_name = os.path.join(folder_path, "loan_applications.xlsx")

# Create the Excel file if it doesn't exist
if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "General Information"
        ws.append(["First Name", "Last Name", "Email", "Phone", "Loan Amount", "Loan Term", "Employment Status", "Annual Income"])
        wb.create_sheet("Work History").append(["Current Employer", "Years at Job", "Previous Employer 1", "Years", "Previous Employer 2", "Years", "Previous Employer 3", "Years"])
        wb.create_sheet("Education History").append(["High School", "Graduation Year", "College 1", "Degree 1", "College 2", "Degree 2", "College 3", "Degree 3"])
        wb.save(file_name)
else:
        wb = openpyxl.load_workbook(file_name)
        sheets_needed = {
            "General Information": ["First Name", "Last Name", "Email", "Phone", "Loan Amount", "Loan Term", "Employment Status", "Annual Income"],
            "Work History": ["Current Employer", "Years at Job", "Previous Employer 1", "Years", "Previous Employer 2", "Years", "Previous Employer 3", "Years"],
            "Education History": ["High School", "Graduation Year", "College 1", "Degree 1", "College 2", "Degree 2", "College 3", "Degree 3"]
        }
        for sheet_name, headers in sheets_needed.items():
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name).append(headers)
        wb.save(file_name)

# Function to validate email
def is_valid_email(email):
    pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.match(pattern, email)

# Function to validate phone number (eg., 123-456-7891)
def is_valid_phone(phone):
    pattern = r"^\d{3}-\d{3}-\d{4}$"  
    return re.match(pattern, phone)

# Function to validate numeric fields (Loan Amount and Annual Income)
def is_valid_number(value):
    return value.isdigit()

# Function to save data to Excel
def save_data():
    first_name = entry_first_name.get().strip()
    last_name = entry_last_name.get().strip()
    email = entry_email.get().strip()
    phone = entry_phone.get().strip()
    loan_amount = entry_loan_amount.get().strip()
    loan_term = loan_term_var.get()
    employment_status = employment_var.get()
    annual_income = entry_income.get().strip()
    current_employer = entry_current_employer.get()

    # Validate last name
    if len(last_name) > 35:
        messagebox.showerror("Error", "Last Name cannot exceed 35 characters.")
        return

    # Validate email
    if not is_valid_email(email):
        messagebox.showerror("Error", "Please enter a valid email address.")
        return

    # Validate phone number
    if not is_valid_phone(phone):
        messagebox.showerror("Error", "Please enter a valid phone number (321-456-7869).")
        return

    # Validate Loan Amount and Annual Income
    if not is_valid_number(loan_amount):
        messagebox.showerror("Error", "Loan Amount must be a numeric value.")
        return

    if not is_valid_number(annual_income):
        messagebox.showerror("Error", "Annual Income must be a numeric value.")
        return
    
    if not is_valid_number(entry_years_current):
        messagebox.showerror("Error", "Years Worked must be a numeric value.")
        return
    
    if loan_term == "Select One":
        messagebox.showerror("Error", "Please select a valid loan term.")
        return
    
    if employment_var == "Select One":
        messagebox.showerror("Error", "Please select a valid employment option.")
        return

def save_data():
    # General Info
    first_name = entry_first_name.get().strip()
    last_name = entry_last_name.get().strip()
    email = entry_email.get().strip()
    phone = entry_phone.get().strip()
    loan_amount = entry_loan_amount.get().strip()
    loan_term = loan_term_var.get()
    employment_status = employment_var.get()
    annual_income = entry_income.get().strip()

    # Work History
    current_employer = entry_current_employer.get().strip()
    years_current = entry_years_current.get().strip()
    prev_employers = []
    for i in range(2, 5):
        emp = frame2.grid_slaves(row=i, column=1)[0].get().strip()
        yrs = frame2.grid_slaves(row=i, column=3)[0].get().strip()
        prev_employers.append((emp, yrs))

    # Education History
    high_school = entry_high_school_name.get().strip()
    grad_year = entry_graduation_year.get().strip()
    college_1 = entry_college_name.get().strip()
    degree_1 = entry_degree_earned.get().strip()
    college_2 = frame3.grid_slaves(row=3, column=1)[0].get().strip()
    degree_2 = frame3.grid_slaves(row=3, column=3)[0].get().strip()
    college_3 = frame3.grid_slaves(row=4, column=1)[0].get().strip()
    degree_3 = frame3.grid_slaves(row=4, column=3)[0].get().strip()

    # --- Validations ---
    if len(last_name) > 35:
        messagebox.showerror("Error", "Last Name cannot exceed 35 characters.")
        return

    if not is_valid_email(email):
        messagebox.showerror("Error", "Please enter a valid email address.")
        return

    if not is_valid_phone(phone):
        messagebox.showerror("Error", "Please enter a valid phone number (321-456-7869).")
        return

    if not is_valid_number(loan_amount):
        messagebox.showerror("Error", "Loan Amount must be a numeric value.")
        return

    if not is_valid_number(annual_income):
        messagebox.showerror("Error", "Annual Income must be a numeric value.")
        return

    if not is_valid_number(years_current):
        messagebox.showerror("Error", "Years Worked must be a numeric value.")
        return

    if loan_term == "Select One":
        messagebox.showerror("Error", "Please select a valid loan term.")
        return

    if employment_status == "Select One":
        messagebox.showerror("Error", "Please select a valid employment option.")
        return

    for emp, yrs in prev_employers:
        if yrs and not is_valid_number(yrs):
            messagebox.showerror("Error", "Years at previous jobs must be numeric.")
            return

    # --- Create or Load Excel File ---
    try:
        wb = openpyxl.load_workbook(file_name)
        ws1 = wb["General Information"]
        ws2 = wb["Work History"]
        ws3 = wb["Education History"]

        # Save General Info
        ws1.append([first_name, last_name, email, phone, loan_amount, loan_term, employment_status, annual_income])

        # Save Work History
        work_row = [current_employer, years_current]
        for emp, yrs in prev_employers:
            work_row.extend([emp, yrs])
        ws2.append(work_row)

        # Save Education History
        ws3.append([high_school, grad_year, college_1, degree_1, college_2, degree_2, college_3, degree_3])
        wb.save(file_name)
        wb.close()
        messagebox.showinfo("Success", "Application Submitted Successfully!")

        # Clear entries
        for entry in [entry_first_name, entry_last_name, entry_email, entry_phone,
                      entry_loan_amount, entry_income, entry_current_employer,
                      entry_years_current, entry_high_school_name, entry_graduation_year,
                      entry_college_name, entry_degree_earned]:
            entry.delete(0, tk.END)

        for row in range(2, 5):
            frame2.grid_slaves(row=row, column=1)[0].delete(0, tk.END)
            frame2.grid_slaves(row=row, column=3)[0].delete(0, tk.END)

        frame3.grid_slaves(row=3, column=1)[0].delete(0, tk.END)
        frame3.grid_slaves(row=3, column=3)[0].delete(0, tk.END)
        frame3.grid_slaves(row=4, column=1)[0].delete(0, tk.END)
        frame3.grid_slaves(row=4, column=3)[0].delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data: {e}")

# Function to switch frames
def show_frame(frame):
    frame.tkraise()

# Create main application window
root = tk.Tk()
root.title("Acme Home Mortgage Application")
root.geometry("850x400")

# Create container for frames
container = tk.Frame(root)
container.place(relx=0.5, rely=0.1, anchor="n")  

# Frame 1
frame1 = tk.Frame(container)
frame1.grid(row=0, column=0, sticky="nsew")

title_label = tk.Label(frame1, text="Acme Home Mortgage Application", font=("Arial", 14, "bold"))
title_label.grid(row=0, column=0, columnspan=4, pady=10)

tk.Label(frame1, text="First Name:").grid(row=1, column=0, sticky="e", padx=5, pady=2)
entry_first_name = tk.Entry(frame1, width=25)
entry_first_name.grid(row=1, column=1, sticky="w")

tk.Label(frame1, text="Last Name:").grid(row=1, column=2, sticky="e", padx=5, pady=2)
entry_last_name = tk.Entry(frame1, width=35)
entry_last_name.grid(row=1, column=3, sticky="w")

tk.Label(frame1, text="Email:").grid(row=2, column=0, sticky="e", padx=5, pady=2)
entry_email = tk.Entry(frame1, width=25)
entry_email.grid(row=2, column=1, sticky="w")

tk.Label(frame1, text="Phone:").grid(row=2, column=2, sticky="e", padx=5, pady=2)
entry_phone = tk.Entry(frame1, width=25)
entry_phone.grid(row=2, column=3, sticky="w")

tk.Label(frame1, text="Loan Amount:").grid(row=3, column=0, sticky="e", padx=5, pady=2)
entry_loan_amount = tk.Entry(frame1, width=25)
entry_loan_amount.grid(row=3, column=1, sticky="w")

# Employment Status and Annual Income
tk.Label(frame1, text="Employment Status:").grid(row=4, column=0, sticky="e", padx=10, pady=5)
employment_var = tk.StringVar(value="Select One")
employment_options = ["Select One", "Employed", "Self-Employed", "Unemployed", "Student", "Retired"]
employment_menu = tk.OptionMenu(frame1, employment_var, *employment_options)
employment_menu.grid(row=4, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame1, text="Annual Income:").grid(row=5, column=0, sticky="e", padx=10, pady=5)
entry_income = tk.Entry(frame1, width=25)
entry_income.grid(row=5, column=1, padx=5, pady=5, sticky="w")

# Dropdowns
tk.Label(frame1, text="Loan Term:").grid(row=3, column=2, sticky="e", padx=10, pady=5)
loan_term_var = tk.StringVar(value="Select One")
loan_term_options = ["Select One", "10-year", "15-year", "30-year"]
loan_term_menu = tk.OptionMenu(frame1, loan_term_var, *loan_term_options)
loan_term_menu.grid(row=3, column=3, padx=5, pady=5, sticky="w")

next_button = tk.Button(frame1, text="Next →", bg="blue", fg="white", font=("Arial", 12), command=lambda: show_frame(frame2))
next_button.grid(row=6, column=3, pady=20, sticky="e", padx=10)

# Frame 2
frame2 = tk.Frame(container)
frame2.grid(row=0, column=0, sticky="nsew")

tk.Label(frame2, text="Work History", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=4, pady=10)

tk.Label(frame2, text="Current Employer:").grid(row=1, column=0, sticky="e", padx=10, pady=5)
entry_current_employer = tk.Entry(frame2, width=25, justify="center")
entry_current_employer.grid(row=1, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame2, text="Years at Job:").grid(row=1, column=2, sticky="e", padx=10, pady=5)
entry_years_current = tk.Entry(frame2, width=10, justify="center")
entry_years_current.grid(row=1, column=3, padx=5, pady=5, sticky="w")

for i in range(3):
  tk.Label(frame2, text="Previous Employer:").grid(row=i+2, column=0, sticky="e", padx=10, pady=5)
  tk.Entry(frame2, width=25, justify="center").grid(row=i+2, column=1, padx=5, pady=5, sticky="w")
  tk.Label(frame2, text="Years at Job:").grid(row=i+2, column=2, sticky="e", padx=10, pady=5)
  tk.Entry(frame2, width=10, justify="center").grid(row=i+2, column=3, padx=5, pady=5, sticky="w")

# Frame 3
frame3 = tk.Frame(container)
frame3.grid(row=0, column=0, sticky="nsew")

tk.Label(frame3, text="Education History", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=4, pady=10)

tk.Label(frame3, text="High School Name:").grid(row=1, column=0, sticky="e", padx=10, pady=5)
entry_high_school_name = tk.Entry(frame3, width=25, justify="center")
entry_high_school_name.grid(row=1, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame3, text="Graduation Year:").grid(row=1, column=2, sticky="e", padx=10, pady=5)
entry_graduation_year = tk.Entry(frame3, width=10, justify="center")
entry_graduation_year.grid(row=1, column=3, padx=5, pady=5, sticky="w")

# 1st College Attended
tk.Label(frame3, text="College Attended:").grid(row=2, column=0, sticky="e", padx=10, pady=5)
entry_college_name = tk.Entry(frame3, width=25, justify="center")
entry_college_name.grid(row=2, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame3, text="Degree Earned:").grid(row=2, column=2, sticky="e", padx=10, pady=5)
entry_degree_earned = tk.Entry(frame3, width=15, justify="center")
entry_degree_earned.grid(row=2, column=3, padx=5, pady=5, sticky="w")

# 2nd College Attended
tk.Label(frame3, text="2nd Previous College Attended:").grid(row=3, column=0, sticky="e", padx=10, pady=5)
entry_college_2 = tk.Entry(frame3, width=25, justify="center")
entry_college_2.grid(row=3, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame3, text="Degree Earned:").grid(row=3, column=2, sticky="e", padx=10, pady=5)
entry_degree_2 = tk.Entry(frame3, width=15, justify="center")
entry_degree_2.grid(row=3, column=3, padx=5, pady=5, sticky="w")

# 3rd College Attended
tk.Label(frame3, text="3rd Previous College Attended:").grid(row=4, column=0, sticky="e", padx=10, pady=5)
entry_college_3 = tk.Entry(frame3, width=25, justify="center")
entry_college_3.grid(row=4, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame3, text="Degree Earned:").grid(row=4, column=2, sticky="e", padx=10, pady=5)
entry_degree_3 = tk.Entry(frame3, width=15, justify="center")
entry_degree_3.grid(row=4, column=3, padx=5, pady=5, sticky="w")


# Back and Submit buttons (Frame 3)
button_frame = tk.Frame(frame3)
button_frame.grid(row=5, column=0, columnspan=4, pady=10)

back_button = tk.Button(button_frame, text="← Back", bg="gray", fg="white", font=("Arial", 12), command=lambda: show_frame(frame2))
back_button.pack(side="left", padx=5)

submit_button = tk.Button(button_frame, text="Submit Application", command=save_data, bg="green", fg="white", font=("Arial", 12, "bold"))
submit_button.pack(side="left", padx=5)

# Back and Submit buttons (Frame 2)
button_frame = tk.Frame(frame2)
button_frame.grid(row=5, column=0, columnspan=4, pady=10)

next_button = tk.Button(button_frame, text="Next →", bg="blue", fg="white", font=("Arial", 12), command=lambda: show_frame(frame3))
next_button.pack(side="right", padx=5)

# Back and Submit buttons (Frame 1)
back_button = tk.Button(button_frame, text="← Back", bg="gray", fg="white", font=("Arial", 12), command=lambda: show_frame(frame1))
back_button.pack(side="left", padx=5)

# Set the initial frame
show_frame(frame1)

# Run the Tkinter event loop
root.mainloop()
