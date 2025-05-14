import tkinter as tk
from tkinter import messagebox
import openpyxl
import os
import re 

# Ensure the directory exists
folder_path = r"c:\apps"
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# Excel file path
file_name = os.path.join(folder_path, "loan_applications.xlsx")

# Create the Excel file if it doesn't exist
if not os.path.exists(file_name):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Personal Info"
    ws1.append(["First Name", "Last Name", "Email", "Phone", "Loan Amount", "Loan Term", "Employment Status", "Annual Income"])
    ws2 = wb.create_sheet(title="Work History")
    ws2.append(["Current Employer", "Years at Job", "Previous Employer 1", "Years at Job", "Previous Employer 2", "Years at Job", "Previous Employer 3", "Years at Job"])
    wb.save(file_name)
    wb.close()

# Function to switch frames
def show_frame(frame):
    frame.tkraise()

# Function to validate email
def is_valid_email(email):
    pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.match(pattern, email)

# Function to validate phone number
def is_valid_phone(phone):
    pattern = r"^\d{3}-\d{3}-\d{4}$"  
    return re.match(pattern, phone)

# Function to save data
def save_data():
    try:
        wb = openpyxl.load_workbook(file_name)
        ws1 = wb["Personal Info"]
        ws2 = wb["Work History"]

        # Save personal info
        ws1.append([
            entry_first_name.get().strip(),
            entry_last_name.get().strip(),
            entry_email.get().strip(),
            entry_phone.get().strip(),
            entry_loan_amount.get().strip(),
            loan_term_var.get(),
            employment_var.get(),
            entry_income.get().strip()
        ])
        
        # Save work history
        ws2.append([
            entry_current_employer.get().strip(),
            entry_current_years.get().strip(),
            entry_prev_employer1.get().strip(),
            entry_prev_years1.get().strip(),
            entry_prev_employer2.get().strip(),
            entry_prev_years2.get().strip(),
            entry_prev_employer3.get().strip(),
            entry_prev_years3.get().strip()
        ])
        
        wb.save(file_name)
        wb.close()
        messagebox.showinfo("Success", "Application Submitted Successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data: {e}")

# Create main application window
root = tk.Tk()
root.title("Acme Home Mortgage Application")
root.geometry("850x400")  

# Create container frame
container = tk.Frame(root)
container.pack(fill="both", expand=True)

# Page 1: Personal Information
page1 = tk.Frame(container)
page1.pack(fill="both", expand=True)

# Personal Info Labels and Entries
tk.Label(page1, text="Personal Information", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=5)

tk.Label(page1, text="First Name:").grid(row=1, column=0, sticky="e", padx=5, pady=2)
entry_first_name = tk.Entry(page1, width=25)
entry_first_name.grid(row=1, column=1, sticky="w")

tk.Label(page1, text="Last Name:").grid(row=2, column=0, sticky="e", padx=5, pady=2)
entry_last_name = tk.Entry(page1, width=25)
entry_last_name.grid(row=2, column=1, sticky="w")

tk.Label(page1, text="Email:").grid(row=3, column=0, sticky="e", padx=5, pady=2)
entry_email = tk.Entry(page1, width=25)
entry_email.grid(row=3, column=1, sticky="w")

tk.Label(page1, text="Phone:").grid(row=4, column=0, sticky="e", padx=5, pady=2)
entry_phone = tk.Entry(page1, width=25)
entry_phone.grid(row=4, column=1, sticky="w")

tk.Label(page1, text="Loan Amount:").grid(row=5, column=0, sticky="e", padx=5, pady=2)
entry_loan_amount = tk.Entry(page1, width=25)
entry_loan_amount.grid(row=5, column=1, sticky="w")

loan_term_var = tk.StringVar(value="15 years")
tk.Label(page1, text="Loan Term:").grid(row=6, column=0, sticky="e", padx=5, pady=2)
tk.OptionMenu(page1, loan_term_var, "15 years", "30 years").grid(row=6, column=1, sticky="w")

employment_var = tk.StringVar(value="Employed")
tk.Label(page1, text="Employment Status:").grid(row=7, column=0, sticky="e", padx=5, pady=2)
tk.OptionMenu(page1, employment_var, "Employed", "Self-Employed", "Unemployed").grid(row=7, column=1, sticky="w")

tk.Label(page1, text="Annual Income:").grid(row=8, column=0, sticky="e", padx=5, pady=2)
entry_income = tk.Entry(page1, width=25)
entry_income.grid(row=8, column=1, sticky="w")

# Navigation Button
tk.Button(page1, text="Next →", command=lambda: show_frame(page2)).grid(row=9, column=1, pady=10)

# Page 2: Work History
page2 = tk.Frame(container)
page2.pack(fill="both", expand=True)

tk.Label(page2, text="Work History", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=4, pady=5)

tk.Label(page2, text="Current Employer:").grid(row=1, column=0, sticky="e", padx=5, pady=2)
entry_current_employer = tk.Entry(page2, width=25)
entry_current_employer.grid(row=1, column=1, sticky="w")

# Additional Work History Fields ...

# Navigation Buttons
tk.Button(page2, text="← Back", command=lambda: show_frame(page1)).grid(row=6, column=0, pady=10)
tk.Button(page2, text="Submit Application", command=save_data, bg="green", fg="white", font=("Arial", 12, "bold")).grid(row=6, column=3, pady=10)

# Show initial frame
show_frame(page1)
root.mainloop()